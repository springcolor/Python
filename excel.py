#!/usr/bin/env python

from openpyxl import Workbook
import time

# book = Workbook()
# sheet = book.active

# sheet['A1'] = 56
# sheet['A2'] = 43
#
# now = time.strftime("%x")
# sheet['A3'] = now

# sheet['A1'] = 1
# sheet.cell(row=2, column=2).value = 2

# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )
#
# for row in rows:
#     sheet.append(row)

# book.save(r"C:\Users\Admin\Desktop\sample.xlsx")

# ********************************************************************

# import openpyxl
#
# book = openpyxl.load_workbook(r"C:\Users\Admin\Desktop\sample.xlsx")
#
# sheet = book.active
#
# a1 = sheet['A1']
# a2 = sheet['A2']
# a3 = sheet.cell(row=3, column=1)
#
# print(a1.value)
# print(a2.value)
# print(a3.value)

# ********************************************************************

# import openpyxl
#
# book = openpyxl.load_workbook(r"C:\Users\Admin\Desktop\sample.xlsx")
#
# sheet = book.active
#
# cells = sheet['A1': 'B3']
#
# for c1, c2 in cells:
#     print("{0:8} {1:8}".format(c1.value, c2.value))

# ********************************************************************

# from openpyxl import Workbook
#
# book = Workbook()
# sheet = book.active
#
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )
#
# for row in rows:
#     sheet.append(row)
#
# for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
#     for cell in row:
#         print(cell.value, end=" ")
#     print()
#
# book.save(r'C:\Users\Admin\Desktop\sample.xlsx')

# ********************************************************************

# from openpyxl import Workbook
#
# book = Workbook()
# sheet = book.active
#
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )
#
# for row in rows:
#     sheet.append(row)
#
# for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
#     for cell in row:
#         print(cell.value, end=" ")
#     print()
#
# book.save(r'C:\Users\Admin\Desktop\sample.xlsx')

# ********************************************************************

# import openpyxl
# import statistics as stats
#
# book = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\sample.xlsx', data_only=True)
#
# sheet = book.active
#
# rows = sheet.rows
#
# values = []
#
# for row in rows:
#     for cell in row:
#         values.append(cell.value)
#
# print("Number of values: {0}".format(len(values)))
# print("Sum of values: {0}".format(sum(values)))
# print("Minimum value: {0}".format(min(values)))
# print("Maximum value: {0}".format(max(values)))
# print("Mean: {0}".format(stats.mean(values)))
# print("Median: {0}".format(stats.median(values)))
# print("Standard deviation: {0}".format(stats.stdev(values)))
# print("Variance: {0}".format(stats.variance(values)))

# ********************************************************************

# from openpyxl import Workbook
#
# wb = Workbook()
# sheet = wb.active
#
# data = [
#     ['Item', 'Colour'],
#     ['pen', 'brown'],
#     ['book', 'black'],
#     ['plate', 'white'],
#     ['chair', 'brown'],
#     ['coin', 'gold'],
#     ['bed', 'brown'],
#     ['notebook', 'white'],
# ]
#
# for r in data:
#     sheet.append(r)
#
# sheet.auto_filter.ref = 'A1:B8'
# sheet.auto_filter.add_filter_column(1, ['brown', 'white'])
# sheet.auto_filter.add_sort_condition('B2:B8')
#
# wb.save(r'C:\Users\Admin\Desktop\sample.xlsx')

# ********************************************************************

# from openpyxl import Workbook
# from openpyxl.chart import (
#     Reference,
#     Series,
#     BarChart
# )
#
# book = Workbook()
# sheet = book.active
#
# rows = [
#     ("USA", 46),
#     ("China", 38),
#     ("UK", 29),
#     ("Russia", 22),
#     ("South Korea", 13),
#     ("Germany", 11)
# ]
#
# for row in rows:
#     sheet.append(row)
#
# data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=6)
# categs = Reference(sheet, min_col=1, min_row=1, max_row=6)
#
# chart = BarChart()
# chart.add_data(data=data)
# chart.set_categories(categs)
#
# chart.legend = None
# chart.y_axis.majorGridlines = None
# chart.varyColors = True
# chart.title = "Olympic Gold medals in London"
#
# sheet.add_chart(chart, "A8")
#
# book.save(r'C:\Users\Admin\Desktop\sample.xlsx')


# ********************************************************************


# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# # import Pillow
#
# book = Workbook()
# sheet = book.active
#
# img = Image(r'C:\Users\Admin\Desktop\456555.png')
# sheet['A1'] = 'This is Sid'
#
# sheet.add_image(img, 'B2')
#
# book.save(r'C:\Users\Admin\Desktop\sample.xlsx')

# ********************************************************************

# from openpyxl import Workbook
#
# book = Workbook()
# sheet = book.active
#
# rows = (
#     (34, 26),
#     (88, 36),
#     (24, 29),
#     (15, 22),
#     (56, 13),
#     (76, 18)
# )
#
# for row in rows:
#     sheet.append(row)
#
# cell = sheet.cell(row=7, column=2)
# cell.value = "=SUM(A1:B6)"
#
# book.save(r'C:\Users\Admin\Desktop\sample.xlsx')

# ********************************************************************

# import openpyxl
#
# book = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\sample.xlsx')

# print(book.get_sheet_names())
#
# active_sheet = book.active
# print(type(active_sheet))
#
# sheet = book.get_sheet_by_name("March")
# print(sheet.title)

# book.create_sheet("April")
#
# print(book.sheetnames)

# try:
#     sheet1 = book.get_sheet_by_name("January")
#     book.remove_sheet(sheet1)
# except[TypeError]:
#     print("程序发生了数字格式异常、算术异常之一")
#     print("未知异常")
# ws = book['sheet1']
# ws.title = 'gouzi'
# print(book.sheetnames)

# book.create_sheet("January", 0)
# print(book.sheetnames)

import openpyxl

book = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\sample.xlsx')
sheet = book.active
# sheet = book.get_sheet_by_name("March")
sheet.sheet_properties.tabColor = "0072BA"

book.save(r'C:\Users\Admin\Desktop\sample.xlsx')